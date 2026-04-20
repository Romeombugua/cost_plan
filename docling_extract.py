"""
Cost Plan PDF Table Extractor (Docling AI) + NRM Enrichment
============================================================
Uses IBM Docling's TableFormer AI model to extract complex tables from
cost plan PDFs, then enriches each line item with the best-matching
NRM sub-element code using a two-stage ML pipeline:

  1. Sentence-transformer (all-MiniLM-L6-v2) computes semantic similarity
     to find the top-K NRM candidates.
  2. Local LLM via Ollama (llama3.2:1b) re-ranks / verifies the best match
     using construction domain knowledge.

Usage:
    python docling_extract.py <pdf_file_or_folder> [--nrm-db path/to/nrm_db.xlsx]
                                                    [--no-llm]

Dependencies:
    pip install docling openpyxl sentence-transformers numpy requests
"""

import sys
import re
import json
import logging
from pathlib import Path

import numpy as np
import openpyxl
import requests
from docling.document_converter import DocumentConverter
from data.icms_data import ICMS_CATALOGUE
from data.uniclass_data import UNICLASS_CATALOGUE
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Suppress noisy third-party loggers
for noisy in (
    "docling", "transformers", "PIL", "urllib3",
    "huggingface_hub", "sentence_transformers",
):
    logging.getLogger(noisy).setLevel(logging.WARNING)


# ═══════════════════════════════════════════════════════════════════════════
#  STYLING CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=11)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
NRM_FONT = Font(name="Calibri", size=11, color="2E75B6")
NRM_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ICMS_FONT = Font(name="Calibri", size=11, color="375623")
ICMS_FILL = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
UNICLASS_FONT = Font(name="Calibri", size=11, color="7030A0")
UNICLASS_FILL = PatternFill(start_color="F3EAF9", end_color="F3EAF9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Columns that clearly aren't descriptions (skip for NRM matching)
SKIP_PATTERNS = re.compile(
    r"^[\d£$€,.\-\s%]+$"        # purely numeric / currency
    r"|^(sub\s*total|total|incl|excl|n/a|tbd|tbc)\.?$",  # summary labels
    re.IGNORECASE,
)


# ═══════════════════════════════════════════════════════════════════════════
#  NRM MATCHER  (sentence-transformer pipeline)
# ═══════════════════════════════════════════════════════════════════════════

class NRMMatcher:
    """
    Semantic matcher that maps free-text line item descriptions to
    NRM sub-element codes using sentence embeddings + cosine similarity.
    """

    MODEL_NAME = "all-MiniLM-L6-v2"       # ~80 MB, very fast on CPU
    DEFAULT_THRESHOLD = 0.35               # minimum similarity to assign a code

    # Common cost plan terms → NRM code mappings for synonym augmentation.
    # These supplement the ML matching for terms where the NRM database
    # uses very different language than typical cost plans.
    SYNONYMS = {
        "2.5.1": ["External walls", "External wall cladding", "Cladding"],
        "2.5.2": ["External walls below ground"],
        "2.7.1": ["Internal walls", "Internal partitions", "Partitions"],
        "2.1.1": ["Frame", "Steel frame", "Structural frame", "Structural steelwork"],
        "2.1.4": ["Concrete frame", "RC frame"],
        "2.1.5": ["Timber frame"],
        "2.2.1": ["Upper floors", "Suspended floors", "Floor slabs"],
        "2.3.2": ["Roof finishes", "Roof covering"],
        "2.4.1": ["Staircases", "Stairs"],
        "2.6.1": ["Windows", "External windows", "Glazing"],
        "2.6.2": ["External doors", "Windows and external doors"],
        "2.8.1": ["Internal doors"],
        "1.1.1": ["Substructure", "Foundations", "Standard foundations", "Ground floor slab"],
        "1.1.3": ["Ground floor construction", "Lowest floor"],
        "3.2.1": ["Floor finishes", "Floor coverings"],
        "3.3.1": ["Ceiling finishes", "Ceilings"],
        "5.1.1": ["Sanitaryware", "Sanitary fittings", "Sanitary installations"],
        "5.3.1": ["Drainage", "Foul drainage", "Above ground drainage"],
        "5.4.1": ["Water supply", "Plumbing", "Water installations"],
        "5.5.1": ["Heating", "Heat source", "Boilers", "HVAC"],
        "5.6.1": ["Central heating", "Heating installation", "Mechanical services"],
        "5.7.1": ["Ventilation", "Central ventilation", "Ductwork"],
        "5.8.1": ["Electrical services", "Electrical installation", "Electrical mains"],
        "5.8.3": ["Lighting", "Lighting installations"],
        "5.10.1": ["Lift installation", "Lifts", "Elevators", "Lift"],
        "5.11.1": ["Fire fighting", "Fire protection systems"],
        "5.12.1": ["Communication systems", "Data installations", "IT infrastructure"],
        "5.12.2": ["Security systems", "CCTV", "Intruder alarm", "Access control"],
        "5.14.1": ["Builders work in connection", "BWIC", "Builder's work"],
        "8.2.1": ["External paving", "Roads and paths", "Hardstanding", "Hardscape"],
        "8.3.1": ["Landscaping", "Seeding", "Turfing", "Soft landscaping"],
        "8.3.2": ["External planting", "Planting"],
        "8.4.1": ["Fencing", "Fencing and railings", "Boundary fencing"],
        "8.6.1": ["External drainage", "Site drainage", "Surface water drainage"],
        "8.7.1": ["External services", "Incoming services", "Utilities"],
        "0.1.3": ["Site clearance", "Demolition", "Enabling works"],
    }

    def __init__(self, nrm_db_path, threshold=None):
        self.threshold = threshold or self.DEFAULT_THRESHOLD
        self.codes = []          # list of nrm_subelement_code strings
        self.descriptions = []   # text used for embedding
        self.short_names = []    # human-readable short names for display
        self.embeddings = None   # (N, dim) numpy matrix
        self.model = None

        self._load_nrm_db(nrm_db_path)
        self._add_synonyms()
        self._build_index()

    # ------------------------------------------------------------------
    def _load_nrm_db(self, path):
        """Load unique NRM sub-element codes + descriptions from Excel."""
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]

        seen = set()
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            code, desc, defn = row
            if not code or code in seen:
                continue
            seen.add(code)

            # Clean the description: strip leading number
            clean_desc = re.sub(r"^\d+\s+", "", str(desc)).strip()

            self.codes.append(str(code))
            # Short name for display in Excel
            self.short_names.append(clean_desc)
            # For embedding: use description only (not the long definition)
            # so short cost plan queries match well against short NRM names
            self.descriptions.append(clean_desc)

        wb.close()
        log.info("Loaded %d unique NRM sub-element codes", len(self.codes))

    # ------------------------------------------------------------------
    def _add_synonyms(self):
        """Inject synonym entries into the index for better matching."""
        # Build code → short_name lookup
        code_to_name = dict(zip(self.codes, self.short_names))
        added = 0
        for code, aliases in self.SYNONYMS.items():
            display_name = code_to_name.get(code)
            if not display_name:
                continue
            for alias in aliases:
                self.codes.append(code)
                self.short_names.append(display_name)
                self.descriptions.append(alias)
                added += 1
        log.info("Added %d synonym entries to NRM index", added)

    # ------------------------------------------------------------------
    def _build_index(self):
        """Encode all NRM descriptions into an embedding matrix."""
        from sentence_transformers import SentenceTransformer

        log.info("Loading sentence-transformer model '%s' ...", self.MODEL_NAME)
        self.model = SentenceTransformer(self.MODEL_NAME)
        log.info("Encoding NRM descriptions ...")
        self.embeddings = self.model.encode(
            self.descriptions,
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        log.info("NRM embedding index ready  (%d entries)", len(self.codes))

    # ------------------------------------------------------------------
    def match_top_k(self, text, k=5):
        """
        Find the top-K NRM matches for a line item description.

        Returns list of (nrm_code, nrm_desc, score) tuples, sorted by
        descending similarity. Returns empty list if text is skippable.
        """
        if not text or SKIP_PATTERNS.match(text.strip()):
            return []

        query_emb = self.model.encode(
            [text], normalize_embeddings=True, show_progress_bar=False,
        )
        similarities = (query_emb @ self.embeddings.T).flatten()

        # Get top-K indices
        top_indices = np.argsort(similarities)[::-1][:k]
        candidates = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= self.threshold:
                candidates.append((
                    self.codes[idx],
                    self.short_names[idx],
                    f"{score * 100:.1f}%",
                ))
        return candidates

    # ------------------------------------------------------------------
    def match(self, text):
        """Single best match (convenience wrapper around match_top_k)."""
        candidates = self.match_top_k(text, k=1)
        if candidates:
            return candidates[0]
        return None, None, None

    # ------------------------------------------------------------------
    def enrich_dataframe(self, df, llm_verifier=None, icms_matcher=None, uniclass_matcher=None):
        """
        Add NRM, ICMS, and/or Uniclass code columns to a DataFrame.
        Optionally uses an LLM verifier for NRM re-ranking.
        """
        nrm_codes = []
        nrm_descs = []
        nrm_confs = []
        icms_codes = [] if icms_matcher else None
        icms_descs = [] if icms_matcher else None
        uniclass_codes = [] if uniclass_matcher else None
        uniclass_descs = [] if uniclass_matcher else None

        # Find the description column (first column with mostly text values)
        desc_col = self._find_description_column(df)

        for _, row in df.iterrows():
            text = str(row.iloc[desc_col]).strip() if desc_col is not None else ""
            candidates = self.match_top_k(text, k=5)

            if not candidates:
                nrm_codes.append("")
                nrm_descs.append("")
                nrm_confs.append("")
            else:
                # Use LLM to verify/re-rank if available
                if llm_verifier and len(candidates) > 1:
                    code, desc, conf = llm_verifier.verify(text, candidates)
                else:
                    code, desc, conf = candidates[0]

                nrm_codes.append(code or "")
                nrm_descs.append(desc or "")
                nrm_confs.append(conf if conf else "")

            if icms_matcher is not None:
                ic, id_ = icms_matcher.match(text)
                icms_codes.append(ic or "")
                icms_descs.append(id_ or "")

            if uniclass_matcher is not None:
                uc, ud = uniclass_matcher.match(text)
                uniclass_codes.append(uc or "")
                uniclass_descs.append(ud or "")

        df = df.copy()
        df["NRM Code"] = nrm_codes
        df["NRM Description"] = nrm_descs
        df["Match Confidence"] = nrm_confs
        if icms_matcher is not None:
            df["ICMS Code"] = icms_codes
            df["ICMS Description"] = icms_descs
        if uniclass_matcher is not None:
            df["Uniclass Code"] = uniclass_codes
            df["Uniclass Description"] = uniclass_descs
        return df

    # ------------------------------------------------------------------
    @staticmethod
    def _find_description_column(df):
        """Identify which column contains the line item descriptions."""
        best_col = None
        best_score = 0

        for col_idx in range(min(len(df.columns), 5)):  # check first 5 cols
            text_count = 0
            for val in df.iloc[:, col_idx]:
                s = str(val).strip()
                if s and len(s) > 3 and not SKIP_PATTERNS.match(s):
                    text_count += 1
            # Prefer columns with more text entries
            if text_count > best_score:
                best_score = text_count
                best_col = col_idx

        return best_col


# ═══════════════════════════════════════════════════════════════════════════
#  ICMS MATCHER  (semantic matching against ICMS 3rd Edition catalogue)
# ═══════════════════════════════════════════════════════════════════════════

class IcmsMatcher:
    """
    Semantic matcher that maps free-text descriptions to ICMS 3rd Edition
    codes using sentence embeddings.  Accepts a pre-loaded SentenceTransformer
    model so the ~80 MB model weights are only loaded once when used alongside
    NRMMatcher.
    """

    MODEL_NAME = "all-MiniLM-L6-v2"
    DEFAULT_THRESHOLD = 0.30

    def __init__(self, model=None, threshold=None):
        self.threshold = threshold or self.DEFAULT_THRESHOLD
        self.codes = []
        self.descriptions = []
        self.embeddings = None
        self.model = model  # share with NRMMatcher to avoid double load

        self._build_index()

    def _build_index(self):
        for elem in ICMS_CATALOGUE:
            self.codes.append(elem.code)
            self.descriptions.append(elem.description)

        if self.model is None:
            from sentence_transformers import SentenceTransformer
            log.info("Loading sentence-transformer model for ICMS matching ...")
            self.model = SentenceTransformer(self.MODEL_NAME)

        log.info("Encoding ICMS descriptions (%d entries) ...", len(self.codes))
        self.embeddings = self.model.encode(
            self.descriptions, normalize_embeddings=True, show_progress_bar=False,
        )
        log.info("ICMS embedding index ready")

    def match(self, text):
        """Return (icms_code, icms_description) for best match, or (None, None)."""
        if not text or SKIP_PATTERNS.match(text.strip()):
            return None, None
        query_emb = self.model.encode(
            [text], normalize_embeddings=True, show_progress_bar=False,
        )
        sims = (query_emb @ self.embeddings.T).flatten()
        idx = int(np.argmax(sims))
        if float(sims[idx]) >= self.threshold:
            return self.codes[idx], self.descriptions[idx]
        return None, None


# ═══════════════════════════════════════════════════════════════════════════
#  UNICLASS MATCHER  (semantic matching against Uniclass 2015 EF table)
# ═══════════════════════════════════════════════════════════════════════════

class UniclassMatcher:
    """
    Semantic matcher that maps free-text descriptions to Uniclass 2015
    codes (EF — Elements/Functions table by default).
    Accepts a pre-loaded SentenceTransformer model for efficient reuse.
    """

    MODEL_NAME = "all-MiniLM-L6-v2"
    DEFAULT_THRESHOLD = 0.30

    def __init__(self, model=None, threshold=None, table="EF"):
        self.threshold = threshold or self.DEFAULT_THRESHOLD
        self.table = table
        self.codes = []
        self.descriptions = []
        self.embeddings = None
        self.model = model

        self._build_index()

    def _build_index(self):
        for entry in UNICLASS_CATALOGUE:
            if entry.table == self.table:
                self.codes.append(entry.code)
                self.descriptions.append(entry.description)

        if self.model is None:
            from sentence_transformers import SentenceTransformer
            log.info("Loading sentence-transformer model for Uniclass matching ...")
            self.model = SentenceTransformer(self.MODEL_NAME)

        log.info(
            "Encoding Uniclass %s descriptions (%d entries) ...",
            self.table, len(self.codes),
        )
        self.embeddings = self.model.encode(
            self.descriptions, normalize_embeddings=True, show_progress_bar=False,
        )
        log.info("Uniclass %s embedding index ready", self.table)

    def match(self, text):
        """Return (uniclass_code, uniclass_description) for best match, or (None, None)."""
        if not text or SKIP_PATTERNS.match(text.strip()):
            return None, None
        query_emb = self.model.encode(
            [text], normalize_embeddings=True, show_progress_bar=False,
        )
        sims = (query_emb @ self.embeddings.T).flatten()
        idx = int(np.argmax(sims))
        if float(sims[idx]) >= self.threshold:
            return self.codes[idx], self.descriptions[idx]
        return None, None


# ═══════════════════════════════════════════════════════════════════════════
#  LLM VERIFIER  (Ollama + llama3.2:1b)
# ═══════════════════════════════════════════════════════════════════════════

class OllamaLLMVerifier:
    """
    Uses a local Ollama LLM to re-rank / verify the top-K NRM candidates
    produced by the sentence-transformer.  The LLM applies construction
    domain knowledge to pick the single best match.
    """

    DEFAULT_MODEL = "llama3.2:1b"
    DEFAULT_BASE_URL = "http://localhost:11434"

    def __init__(self, model=None, base_url=None):
        self.model = model or self.DEFAULT_MODEL
        self.base_url = (base_url or self.DEFAULT_BASE_URL).rstrip("/")

    # ------------------------------------------------------------------
    def ping(self):
        """Return True if Ollama is reachable and has the model."""
        try:
            resp = requests.get(f"{self.base_url}/api/tags", timeout=5)
            if resp.status_code != 200:
                return False
            models = [m["name"] for m in resp.json().get("models", [])]
            # Accept both "llama3.2:1b" and "llama3.2:1b" with tag suffix
            if not any(self.model in m for m in models):
                log.warning(
                    "Ollama is running but model '%s' not found. "
                    "Available: %s. Pull it with: ollama pull %s",
                    self.model, models, self.model,
                )
                return False
            return True
        except requests.ConnectionError:
            return False

    # ------------------------------------------------------------------
    def _call_ollama(self, prompt):
        """
        Send a prompt to the Ollama /api/generate endpoint.
        Returns the response text or None on failure.
        """
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.1,       # near-deterministic
                "num_predict": 200,       # short answers only
            },
        }
        try:
            resp = requests.post(
                f"{self.base_url}/api/generate",
                json=payload,
                timeout=30,
            )
            resp.raise_for_status()
            return resp.json().get("response", "").strip()
        except Exception as exc:
            log.debug("Ollama call failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    def verify(self, description, candidates):
        """
        Ask the LLM to pick the best NRM code for *description* from
        the list of (code, nrm_desc, score) *candidates*.

        Returns (code, description, confidence) — falls back to the
        top sentence-transformer candidate on any failure.
        """
        fallback = candidates[0]  # best sentence-transformer match

        # Build a numbered option list for the prompt
        options = "\n".join(
            f"  {i+1}. {code} — {desc} (similarity: {score})"
            for i, (code, desc, score) in enumerate(candidates)
        )

        prompt = (
            "You are an expert UK construction cost consultant.\n"
            "A cost plan line item is described as:\n"
            f'  "{description}"\n\n'
            "The following NRM sub-element codes are possible matches "
            "(ranked by semantic similarity):\n"
            f"{options}\n\n"
            "Pick the SINGLE best NRM code for the line item. "
            "Reply with ONLY a JSON object, no other text:\n"
            '{"code": "<nrm_code>", "reason": "<brief reason>"}\n'
        )

        raw = self._call_ollama(prompt)
        if not raw:
            return fallback

        # Parse JSON from the LLM response
        try:
            # The LLM may wrap JSON in markdown fences — strip them
            cleaned = raw.strip().strip("`").strip()
            if cleaned.startswith("json"):
                cleaned = cleaned[4:].strip()
            data = json.loads(cleaned)
            chosen_code = str(data.get("code", "")).strip()
        except (json.JSONDecodeError, AttributeError):
            log.debug("LLM response not parseable as JSON: %s", raw[:200])
            return fallback

        # Map the chosen code back to the candidate list
        for code, desc, score in candidates:
            if code == chosen_code:
                reason = data.get("reason", "")
                log.debug(
                    "LLM verified: '%s' → %s (%s) [%s]",
                    description[:40], code, score, reason,
                )
                return (code, desc, score)

        # LLM picked a code not in the candidates — fall back
        log.debug("LLM chose unknown code '%s'; falling back", chosen_code)
        return fallback


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════

def write_tables_to_excel(tables_data, output_path):
    """
    Write all extracted tables continuously onto a single Excel sheet.
    Tables are stacked vertically with a label row and a blank-row gap.
    """
    LABEL_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")
    GAP_ROWS = 2

    # Column-name sets for per-family styling
    NRM_COLS = {"NRM Code", "NRM Description", "Match Confidence"}
    ICMS_COLS = {"ICMS Code", "ICMS Description"}
    UNICLASS_COLS = {"Uniclass Code", "Uniclass Description"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    if not tables_data:
        ws.cell(row=1, column=1, value="No tables were detected in this PDF.")
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=12, italic=True)
        wb.save(output_path)
        log.info("Saved  %s  (no tables)", output_path)
        return

    current_row = 1

    for t_idx, entry in enumerate(tables_data):
        label = entry["sheet_name"]
        df = entry["dataframe"]
        col_names = list(df.columns)

        # -- Table label row --
        label_cell = ws.cell(
            row=current_row, column=1,
            value=f"Table {t_idx + 1} -- {label}",
        )
        label_cell.font = LABEL_FONT
        current_row += 1

        # -- Header row --
        for c_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        current_row += 1

        # -- Data rows --
        for row_tuple in df.itertuples(index=False):
            for c_idx, value in enumerate(row_tuple, start=1):
                col_name = col_names[c_idx - 1]
                cell = ws.cell(row=current_row, column=c_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT
                if col_name in NRM_COLS:
                    cell.font = NRM_FONT
                    cell.fill = NRM_FILL
                elif col_name in ICMS_COLS:
                    cell.font = ICMS_FONT
                    cell.fill = ICMS_FILL
                elif col_name in UNICLASS_COLS:
                    cell.font = UNICLASS_FONT
                    cell.fill = UNICLASS_FILL
                else:
                    cell.font = CELL_FONT
            current_row += 1

        current_row += GAP_ROWS

    # -- Auto-size columns --
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    log.info("Saved  %s  (%d table(s) on 1 sheet)", output_path, len(tables_data))


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def process_pdf(converter, pdf_path, nrm_matcher=None, llm_verifier=None,
                icms_matcher=None, uniclass_matcher=None):
    """
    Extract every table from a PDF using Docling, optionally enrich
    with NRM, ICMS, and/or Uniclass codes, and write to Excel.
    """
    pdf_path = Path(pdf_path)
    output_path = pdf_path.with_name(pdf_path.stem + "_docling_tables.xlsx")

    log.info("Processing  %s", pdf_path.name)

    result = converter.convert(str(pdf_path))
    doc = result.document

    tables = list(doc.tables)
    log.info("  Found %d table(s)", len(tables))

    tables_data = []
    for t_idx, table in enumerate(tables):
        try:
            df = table.export_to_dataframe(doc=doc)

            # Build a meaningful label
            page_label = ""
            if hasattr(table, "prov") and table.prov:
                try:
                    page_no = table.prov[0].page_no
                    page_label = f"P{page_no}"
                except (IndexError, AttributeError):
                    pass

            sheet_name = f"{page_label}_Table_{t_idx + 1}" if page_label else f"Table_{t_idx + 1}"

            # Enrich with NRM / ICMS / Uniclass codes if matchers are available
            if nrm_matcher:
                df = nrm_matcher.enrich_dataframe(
                    df,
                    llm_verifier=llm_verifier,
                    icms_matcher=icms_matcher,
                    uniclass_matcher=uniclass_matcher,
                )

            log.info(
                "  Table %d: %d rows x %d cols %s",
                t_idx + 1, len(df), len(df.columns),
                f"(page {page_label})" if page_label else "",
            )

            tables_data.append({
                "sheet_name": sheet_name,
                "dataframe": df,
            })
        except Exception as exc:
            log.warning("  Table %d: export failed - %s", t_idx + 1, exc)

    write_tables_to_excel(tables_data, output_path)
    return output_path


def _find_nrm_db(target_path):
    """Auto-discover nrm_db.xlsx near the target path."""
    search_dirs = [Path(target_path)]
    if search_dirs[0].is_file():
        search_dirs = [search_dirs[0].parent]
    search_dirs.append(Path.cwd())

    for d in search_dirs:
        candidate = d / "nrm_db.xlsx"
        if candidate.exists():
            return candidate
    return None


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Error: please provide a PDF file or folder path.")
        sys.exit(1)

    # Parse arguments
    target = None
    nrm_db_path = None
    skip_llm = False
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--nrm-db" and i + 1 < len(args):
            nrm_db_path = Path(args[i + 1])
            i += 2
        elif args[i] == "--no-llm":
            skip_llm = True
            i += 1
        else:
            target = Path(args[i])
            i += 1

    if not target:
        print("Error: please provide a PDF file or folder path.")
        sys.exit(1)

    if target.is_file() and target.suffix.lower() == ".pdf":
        pdf_files = [target]
    elif target.is_dir():
        pdf_files = sorted(target.glob("*.pdf"))
        if not pdf_files:
            log.error("No PDF files found in %s", target)
            sys.exit(1)
        log.info("Found %d PDF(s) in %s", len(pdf_files), target)
    else:
        log.error("Path is not a valid PDF file or directory: %s", target)
        sys.exit(1)

    # Auto-discover NRM database if not explicitly provided
    if not nrm_db_path:
        nrm_db_path = _find_nrm_db(target)

    # Initialise NRM matcher
    nrm_matcher = None
    if nrm_db_path and nrm_db_path.exists():
        log.info("NRM database found: %s", nrm_db_path)
        nrm_matcher = NRMMatcher(nrm_db_path)
    else:
        log.warning("NRM database not found — skipping NRM enrichment.")

    # Initialise ICMS and Uniclass matchers (share model to avoid double load)
    icms_matcher = None
    uniclass_matcher = None
    if nrm_matcher:
        log.info("Building ICMS matcher ...")
        icms_matcher = IcmsMatcher(model=nrm_matcher.model)
        log.info("Building Uniclass matcher ...")
        uniclass_matcher = UniclassMatcher(model=nrm_matcher.model)

    # Initialise LLM verifier (Ollama)
    llm_verifier = None
    if nrm_matcher and not skip_llm:
        verifier = OllamaLLMVerifier()
        if verifier.ping():
            llm_verifier = verifier
            log.info(
                "Ollama LLM verifier ready  (model: %s)",
                verifier.model,
            )
        else:
            log.warning(
                "Ollama not available — LLM verification disabled. "
                "Run 'ollama serve' and 'ollama pull %s' to enable.",
                verifier.model,
            )
    elif skip_llm:
        log.info("LLM verification skipped (--no-llm flag).")

    # Initialise Docling converter
    log.info("Initialising Docling AI models (this may take a moment on first run)...")
    converter = DocumentConverter()
    log.info("Models loaded.")

    results = []
    for pdf_file in pdf_files:
        try:
            out = process_pdf(
                converter, pdf_file, nrm_matcher, llm_verifier,
                icms_matcher, uniclass_matcher,
            )
            results.append((pdf_file.name, out, None))
        except Exception as exc:
            log.error("FAILED  %s : %s", pdf_file.name, exc)
            results.append((pdf_file.name, None, str(exc)))

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for name, out, err in results:
        if err:
            print(f"  [FAIL]  {name}  ->  ERROR: {err}")
        else:
            print(f"  [OK]    {name}  ->  {out.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
