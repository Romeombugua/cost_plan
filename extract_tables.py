"""
Cost Plan PDF Table Extractor
==============================
Extracts tables from cost plan PDFs (text-based or scanned) and writes
them into Excel workbooks.

Usage:
    python extract_tables.py <pdf_file_or_folder>

Dependencies:
    pip install pdfplumber openpyxl easyocr pdf2image
"""

import sys
import os
import re
import logging
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCANNED_PAGE_CHAR_THRESHOLD = 50  # pages with fewer chars are treated as scanned
ROW_CLUSTER_TOLERANCE = 10        # pixels – OCR boxes within this Y-range form one row


# ═══════════════════════════════════════════════════════════════════════════
#  TEXT-BASED EXTRACTION  (pdfplumber)
# ═══════════════════════════════════════════════════════════════════════════

def extract_tables_from_text_page(page):
    """Extract all tables from a text-based PDF page using pdfplumber."""
    table_settings = {
        "vertical_strategy": "lines_strict",
        "horizontal_strategy": "lines_strict",
        "snap_y_tolerance": 5,
        "snap_x_tolerance": 5,
        "join_y_tolerance": 5,
        "join_x_tolerance": 5,
    }

    tables = page.extract_tables(table_settings)

    # If strict strategy finds nothing, fall back to the default strategy
    if not tables:
        tables = page.extract_tables()

    cleaned_tables = []
    for table in tables:
        cleaned = clean_table(table)
        if cleaned:
            cleaned_tables.append(cleaned)

    return cleaned_tables


def clean_table(table):
    """Clean a raw table: strip whitespace, replace None, drop empty rows."""
    if not table:
        return []

    cleaned = []
    for row in table:
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append("")
            else:
                # Collapse internal newlines into spaces, strip edges
                text = re.sub(r"\s+", " ", str(cell)).strip()
                cleaned_row.append(text)
        # Keep the row if it has at least one non-empty cell
        if any(c for c in cleaned_row):
            cleaned.append(cleaned_row)

    return cleaned


# ═══════════════════════════════════════════════════════════════════════════
#  OCR-BASED EXTRACTION  (EasyOCR + pypdfium2)
# ═══════════════════════════════════════════════════════════════════════════

# Lazy-loaded so the heavy torch import only happens when needed.
_ocr_reader = None


def _get_ocr_reader():
    """Return a cached EasyOCR reader instance (lazy-loaded)."""
    global _ocr_reader
    if _ocr_reader is None:
        log.info("Initialising EasyOCR reader (first scanned page) …")
        import easyocr
        _ocr_reader = easyocr.Reader(["en"], gpu=False)
    return _ocr_reader


def page_to_image(page):
    """Render a pdfplumber page to a PIL Image via pypdfium2."""
    # pdfplumber wraps pypdfium2, so .to_image() is available
    img_obj = page.to_image(resolution=300)
    return img_obj.original  # PIL.Image


def extract_tables_from_scanned_page(page):
    """Run EasyOCR on a rendered page image and reconstruct tables."""
    import numpy as np

    reader = _get_ocr_reader()
    pil_img = page_to_image(page)
    img_array = np.array(pil_img)

    results = reader.readtext(img_array)
    if not results:
        return []

    # results = list of (bbox, text, confidence)
    # bbox = [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
    entries = []
    for bbox, text, conf in results:
        xs = [pt[0] for pt in bbox]
        ys = [pt[1] for pt in bbox]
        entries.append({
            "text": text.strip(),
            "x_min": min(xs),
            "x_max": max(xs),
            "y_min": min(ys),
            "y_max": max(ys),
            "y_center": (min(ys) + max(ys)) / 2,
            "x_center": (min(xs) + max(xs)) / 2,
        })

    # --- Cluster into rows by y_center ---
    entries.sort(key=lambda e: e["y_center"])
    rows = []
    current_row = [entries[0]]
    for entry in entries[1:]:
        if abs(entry["y_center"] - current_row[0]["y_center"]) <= ROW_CLUSTER_TOLERANCE:
            current_row.append(entry)
        else:
            rows.append(current_row)
            current_row = [entry]
    rows.append(current_row)

    # Sort cells within each row by x_center
    for row in rows:
        row.sort(key=lambda e: e["x_center"])

    # --- Determine column grid ---
    # Collect all x-centre values and cluster into columns
    all_x = sorted(e["x_center"] for e in entries)
    col_centers = _cluster_values(all_x, tolerance=40)
    num_cols = len(col_centers)

    # Assign each cell to the nearest column
    def nearest_col(x_center):
        return min(range(num_cols), key=lambda i: abs(col_centers[i] - x_center))

    table = []
    for row_entries in rows:
        row_cells = [""] * num_cols
        for entry in row_entries:
            col_idx = nearest_col(entry["x_center"])
            existing = row_cells[col_idx]
            if existing:
                row_cells[col_idx] = existing + " " + entry["text"]
            else:
                row_cells[col_idx] = entry["text"]
        table.append(row_cells)

    cleaned = clean_table(table)
    return [cleaned] if cleaned else []


def _cluster_values(values, tolerance):
    """Cluster a sorted list of numeric values, returning cluster centres."""
    if not values:
        return []
    clusters = [[values[0]]]
    for v in values[1:]:
        if v - clusters[-1][-1] <= tolerance:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return [sum(c) / len(c) for c in clusters]


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=11)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_tables_to_excel(pdf_tables, output_path):
    """
    Write extracted tables to an Excel workbook.

    Parameters
    ----------
    pdf_tables : list of dict
        Each dict has keys: "page" (int), "tables" (list of list-of-lists),
        "source" ("text" | "ocr").
    output_path : str | Path
        Destination .xlsx path.
    """
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    sheet_idx = 0
    for page_info in pdf_tables:
        page_num = page_info["page"]
        source = page_info["source"]
        tables = page_info["tables"]

        for t_idx, table in enumerate(tables):
            sheet_idx += 1
            suffix = f"_T{t_idx + 1}" if len(tables) > 1 else ""
            sheet_name = f"Page_{page_num + 1}{suffix}"
            # Excel sheet names max 31 chars
            sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            # Determine if first row looks like a header
            is_header_row = True

            for r_idx, row in enumerate(table):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.border = THIN_BORDER

                    if r_idx == 0 and is_header_row:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = HEADER_ALIGNMENT
                    else:
                        cell.font = CELL_FONT
                        cell.alignment = CELL_ALIGNMENT

            # Auto-size columns
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                adjusted = min(max_len + 4, 50)
                ws.column_dimensions[col_letter].width = adjusted

            # Add a small note about extraction source
            note_row = len(table) + 2
            note_cell = ws.cell(
                row=note_row, column=1,
                value=f"[Extracted via {'OCR (EasyOCR)' if source == 'ocr' else 'pdfplumber'}]",
            )
            note_cell.font = Font(name="Calibri", size=9, italic=True, color="888888")

    if sheet_idx == 0:
        ws = wb.create_sheet(title="No Tables Found")
        ws.cell(row=1, column=1, value="No tables were detected in this PDF.")
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=12, italic=True)

    wb.save(output_path)
    log.info("Saved  %s  (%d sheet%s)", output_path, sheet_idx, "s" if sheet_idx != 1 else "")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def process_pdf(pdf_path):
    """
    Extract every table from a PDF and write to an Excel workbook.

    Returns the output Excel path.
    """
    pdf_path = Path(pdf_path)
    output_path = pdf_path.with_name(pdf_path.stem + "_tables.xlsx")

    log.info("Processing  %s", pdf_path.name)

    pdf = pdfplumber.open(str(pdf_path))
    all_page_tables = []

    for page_idx, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        is_scanned = len(text.strip()) < SCANNED_PAGE_CHAR_THRESHOLD

        if is_scanned:
            log.info("  Page %d → scanned (OCR)", page_idx + 1)
            tables = extract_tables_from_scanned_page(page)
            source = "ocr"
        else:
            tables = extract_tables_from_text_page(page)
            source = "text"

        if tables:
            log.info(
                "  Page %d → %d table(s) via %s",
                page_idx + 1, len(tables), source,
            )
            all_page_tables.append({
                "page": page_idx,
                "tables": tables,
                "source": source,
            })
        else:
            log.info("  Page %d → no tables found", page_idx + 1)

    pdf.close()

    write_tables_to_excel(all_page_tables, output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Error: please provide a PDF file or folder path.")
        sys.exit(1)

    target = Path(sys.argv[1])

    if target.is_file() and target.suffix.lower() == ".pdf":
        pdf_files = [target]
    elif target.is_dir():
        pdf_files = sorted(target.glob("*.pdf"))
        if not pdf_files:
            log.error("No PDF files found in %s", target)
            sys.exit(1)
        log.info("Found %d PDF(s) in %s", len(pdf_files), target)
    else:
        log.error("Path is not a valid PDF file or directory: %s", target)
        sys.exit(1)

    results = []
    for pdf_file in pdf_files:
        try:
            out = process_pdf(pdf_file)
            results.append((pdf_file.name, out, None))
        except Exception as exc:
            log.error("FAILED  %s : %s", pdf_file.name, exc)
            results.append((pdf_file.name, None, str(exc)))

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for name, out, err in results:
        if err:
            print(f"  [FAIL]  {name}  ->  ERROR: {err}")
        else:
            print(f"  [OK]    {name}  ->  {out.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
