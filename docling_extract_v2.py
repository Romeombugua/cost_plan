"""
Cost Plan PDF Table Extractor (Docling AI) + NRM Enrichment - V2
=================================================================
Accuracy-enhanced version. Key improvements:
  1. Docling ACCURATE mode + cell matching control
  2. Context-aware NRM matching (section headers + adjacent rows)
  3. Improved skip logic for summary/fee/total rows
  4. Optional cross-encoder re-ranking stage
  5. Richer NRM embeddings (description + definition combined)
  6. Externalised synonym dictionary (JSON file)

Usage:
    python docling_extract_v2.py <pdf_file_or_folder>
        [--nrm-db path/to/nrm_db.xlsx]
        [--synonyms path/to/synonyms.json]
        [--no-llm]
        [--use-cross-encoder]
"""

import sys, re, json, logging
from pathlib import Path
import numpy as np
import openpyxl
import requests
from docling.document_converter import DocumentConverter, PdfFormatOption
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions, TableFormerMode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)
for _n in ("docling","transformers","PIL","urllib3","huggingface_hub","sentence_transformers"):
    logging.getLogger(_n).setLevel(logging.WARNING)

# ── Styling ──
HEADER_FONT  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT    = Font(name="Calibri", size=11)
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
NRM_FONT     = Font(name="Calibri", size=11, color="2E75B6")
NRM_FILL     = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
SKIP_FONT    = Font(name="Calibri", size=11, italic=True, color="999999")
THIN_BORDER  = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))

# ── V2: Improved skip / filter logic ──
NUMERIC_ONLY = re.compile(r"^[\d\s\u00a3$,.\-\+%()]+$")
SUMMARY_RE = re.compile(
    r"(?:^|\b)("
    r"sub[\s\-]?total|total|grand total|"
    r"measured work|estimated.*(?:building|current|cost)|"
    r"preliminaries|prelims|"
    r"contingency|contingencies|"
    r"overheads?\s+and\s+profit|oh\s*[&+]\s*p|"
    r"contractor.?s?\s+(?:profit|preliminaries|overheads?)|"
    r"main\s+contractor|"
    r"design\s+(?:development|team)\s+(?:contingency|fees?)|"
    r"project.*(?:fees?|team)|inflation|insurances?|"
    r"provisional\s+sums?|"
    r"construction\s+cost|building\s+cost|"
    r"on\s+costs|"
    r"n/?a|tbd|tbc|excluded"
    r")(?:\b|$)", re.IGNORECASE)

NRM_SECTION_HEADERS = {
    "facilitating works": "0",
    "substructure": "1", "superstructure": "2",
    "internal finishes": "3",
    "fittings, furnishings and equipment": "4",
    "prefabricated buildings": "6", "complete building": "6",
    "work to existing buildings": "7", "external works": "8",
}

# These require exact match (after stripping leading numbers)
# to avoid false positives like "Mechanical Services" matching "Services"
NRM_SECTION_HEADERS_EXACT = {
    "services": "5",
    "fittings": "4",
    "facilitating": "0",
}

def should_skip_row(text):
    if not text or len(text.strip()) < 3:
        return True
    t = text.strip()
    if NUMERIC_ONLY.match(t):
        return True
    if SUMMARY_RE.search(t):
        return True
    return False

def detect_section_header(text):
    """Return the NRM group number if text is a section header, else None."""
    if not text:
        return None
    clean = re.sub(r"^\d+\.?\s*", "", text.strip().lower()).strip()
    # First check substring-safe headers (multi-word, unambiguous)
    for header, group in NRM_SECTION_HEADERS.items():
        if header in clean:
            return group
    # Then check exact-match-only headers (single words that could be substrings)
    for header, group in NRM_SECTION_HEADERS_EXACT.items():
        if clean == header:
            return group
    return None


# ══════════════════════════════════════════════════════════════════════
#  NRM MATCHER V2
# ══════════════════════════════════════════════════════════════════════
class NRMMatcher:
    MODEL_NAME = "all-MiniLM-L6-v2"
    DEFAULT_THRESHOLD = 0.30

    DEFAULT_SYNONYMS = {
        "2.5.1": ["External walls","External wall cladding","Cladding","External enclosing walls","Curtain walling"],
        "2.5.2": ["External walls below ground"],
        "2.7.1": ["Internal walls","Internal partitions","Partitions"],
        "2.1.1": ["Frame","Steel frame","Structural frame","Structural steelwork"],
        "2.1.4": ["Concrete frame","RC frame"], "2.1.5": ["Timber frame"],
        "2.2.1": ["Upper floors","Suspended floors","Floor slabs"],
        "2.3.1": ["Roof structure","Roof"], "2.3.2": ["Roof finishes","Roof covering"],
        "2.4.1": ["Staircases","Stairs","Stairs and ramps"],
        "2.6.1": ["Windows","External windows","Glazing"],
        "2.6.2": ["External doors","Windows and external doors"],
        "2.8.1": ["Internal doors"],
        "1.1.1": ["Substructure","Foundations","Standard foundations","Ground floor slab"],
        "1.1.3": ["Ground floor construction","Lowest floor"],
        "3.1.1": ["Wall finishes"], "3.2.1": ["Floor finishes","Floor coverings"],
        "3.3.1": ["Ceiling finishes","Ceilings"],
        "4.1.1": ["Fittings","Fittings and fixtures","FF&E"],
        "5.1.1": ["Sanitaryware","Sanitary fittings","Sanitary appliances"],
        "5.3.1": ["Drainage","Foul drainage","Disposal installations","Disposal installation"],
        "5.4.1": ["Water supply","Plumbing","Water installations"],
        "5.5.1": ["Heating","Heat source","Boilers","HVAC"],
        "5.6.1": ["Central heating","Heating installation","Mechanical services"],
        "5.7.1": ["Ventilation","Central ventilation","Ductwork","Ventilation systems"],
        "5.8.1": ["Electrical services","Electrical installation","Electrical mains"],
        "5.8.3": ["Lighting","Lighting installations"],
        "5.10.1": ["Lift installation","Lifts","Elevators","Lift and conveyor installations"],
        "5.11.1": ["Fire fighting","Fire protection","Fire and lightning protection"],
        "5.12.1": ["Communication systems","Data installations","IT infrastructure"],
        "5.12.2": ["Security systems","CCTV","Intruder alarm","Access control"],
        "5.14.1": ["Builders work in connection","BWIC","Builder's work"],
        "7.1.1": ["Demolition and alteration","Minor demolition","Demolition","Strip out"],
        "8.1.2": ["Site works","Site preparation","Enabling works","Preparatory groundworks"],
        "8.2.1": ["External paving","Roads and paths","Hardstanding","Hardscape"],
        "8.3.1": ["Landscaping","Seeding","Turfing","Soft landscaping"],
        "8.3.2": ["External planting","Planting"],
        "8.4.1": ["Fencing","Fencing and railings","Boundary fencing"],
        "8.6.1": ["External drainage","Site drainage","Surface water drainage"],
        "8.7.1": ["External services","Incoming services","Utilities"],
        "0.1.3": ["Eradication of plant growth","Japanese knotweed"],
        "0.4.1": ["Specialist groundworks","Piling","Ground improvement"],
    }

    def __init__(self, nrm_db_path, synonyms_path=None, threshold=None):
        self.threshold = threshold or self.DEFAULT_THRESHOLD
        self.codes, self.descriptions, self.short_names = [], [], []
        self.embeddings = None
        self.model = None

        self._load_nrm_db(nrm_db_path)
        self._load_synonyms(synonyms_path)
        self._build_index()

    def _load_nrm_db(self, path):
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        seen = set()
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            code, desc, defn = row
            if not code or code in seen:
                continue
            seen.add(code)
            clean_desc = re.sub(r"^\d+\s+", "", str(desc)).strip()
            clean_defn = str(defn).strip() if defn else ""
            self.codes.append(str(code))
            self.short_names.append(clean_desc)
            # V2: richer embedding text = description + definition
            if clean_defn and clean_defn.lower() != "none":
                self.descriptions.append(f"{clean_desc}. {clean_defn}")
            else:
                self.descriptions.append(clean_desc)
        wb.close()
        log.info("Loaded %d unique NRM sub-element codes", len(self.codes))

    def _load_synonyms(self, synonyms_path):
        synonyms = dict(self.DEFAULT_SYNONYMS)
        if synonyms_path and Path(synonyms_path).exists():
            with open(synonyms_path, "r") as f:
                user_synonyms = json.load(f)
            for code, aliases in user_synonyms.items():
                if code in synonyms:
                    synonyms[code] = list(set(synonyms[code] + aliases))
                else:
                    synonyms[code] = aliases
            log.info("Loaded %d user synonym entries from %s", len(user_synonyms), synonyms_path)

        code_to_name = dict(zip(self.codes, self.short_names))
        added = 0
        for code, aliases in synonyms.items():
            display_name = code_to_name.get(code)
            if not display_name:
                continue
            for alias in aliases:
                self.codes.append(code)
                self.short_names.append(display_name)
                self.descriptions.append(alias)
                added += 1
        log.info("Added %d synonym entries to NRM index", added)

    def _build_index(self):
        from sentence_transformers import SentenceTransformer
        log.info("Loading sentence-transformer '%s' ...", self.MODEL_NAME)
        self.model = SentenceTransformer(self.MODEL_NAME)
        log.info("Encoding %d NRM descriptions ...", len(self.descriptions))
        self.embeddings = self.model.encode(
            self.descriptions, normalize_embeddings=True, show_progress_bar=False)
        log.info("NRM embedding index ready")

    def match_top_k(self, text, k=5, context_group=None):
        """
        V2: Find top-K NRM matches. If context_group is provided,
        boost candidates whose NRM group matches the detected section.
        """
        if not text or should_skip_row(text):
            return []

        query_emb = self.model.encode([text], normalize_embeddings=True, show_progress_bar=False)
        similarities = (query_emb @ self.embeddings.T).flatten()

        # V2: Context-based group boosting
        if context_group is not None:
            for i, code in enumerate(self.codes):
                if code.startswith(context_group + "."):
                    similarities[i] += 0.10  # boost same-group candidates

        top_indices = np.argsort(similarities)[::-1][:k]
        candidates = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= self.threshold:
                candidates.append((self.codes[idx], self.short_names[idx], f"{score*100:.1f}%"))
        return candidates

    def match(self, text, context_group=None):
        candidates = self.match_top_k(text, k=1, context_group=context_group)
        return candidates[0] if candidates else (None, None, None)

    def enrich_dataframe(self, df, llm_verifier=None, cross_encoder=None):
        """
        V2: Context-aware enrichment. Tracks the current NRM section
        header as rows are processed, and passes context to matching.
        """
        nrm_codes, nrm_descs, nrm_confs = [], [], []
        desc_col = self._find_description_column(df)
        current_section_group = None

        for row_idx, row in df.iterrows():
            text = str(row.iloc[desc_col]).strip() if desc_col is not None else ""

            # V2: Detect section headers and update context
            detected_group = detect_section_header(text)
            if detected_group is not None:
                current_section_group = detected_group
                log.debug("Section header detected: group %s from '%s'", detected_group, text[:40])

            # V2: Check if this row should be skipped
            if should_skip_row(text):
                nrm_codes.append("")
                nrm_descs.append("")
                nrm_confs.append("")
                continue

            # V2: Build context string from adjacent rows
            context_parts = [text]
            if row_idx > 0 and desc_col is not None:
                prev_text = str(df.iloc[row_idx - 1, desc_col]).strip()
                if prev_text and prev_text.lower() != "none":
                    context_parts.insert(0, f"[prev: {prev_text}]")
            if row_idx < len(df) - 1 and desc_col is not None:
                next_text = str(df.iloc[row_idx + 1, desc_col]).strip()
                if next_text and next_text.lower() != "none":
                    context_parts.append(f"[next: {next_text}]")

            # Use context-enriched text for matching
            context_text = " ".join(context_parts)
            candidates = self.match_top_k(context_text, k=5, context_group=current_section_group)

            if not candidates:
                nrm_codes.append("")
                nrm_descs.append("")
                nrm_confs.append("")
                continue

            # V2: Cross-encoder re-ranking (if available)
            if cross_encoder and len(candidates) > 1:
                code, desc, conf = cross_encoder.rerank(text, candidates)
            elif llm_verifier and len(candidates) > 1:
                code, desc, conf = llm_verifier.verify(text, candidates)
            else:
                code, desc, conf = candidates[0]

            nrm_codes.append(code or "")
            nrm_descs.append(desc or "")
            nrm_confs.append(conf if conf else "")

        df = df.copy()
        df["NRM Code"] = nrm_codes
        df["NRM Description"] = nrm_descs
        df["Match Confidence"] = nrm_confs
        return df

    @staticmethod
    def _find_description_column(df):
        best_col, best_score = None, 0
        for col_idx in range(min(len(df.columns), 5)):
            text_count = 0
            for val in df.iloc[:, col_idx]:
                s = str(val).strip()
                if s and len(s) > 3 and not NUMERIC_ONLY.match(s):
                    text_count += 1
            if text_count > best_score:
                best_score = text_count
                best_col = col_idx
        return best_col


# ══════════════════════════════════════════════════════════════════════
#  V2: CROSS-ENCODER RE-RANKER (faster alternative to LLM)
# ══════════════════════════════════════════════════════════════════════
class CrossEncoderReranker:
    """
    Uses a cross-encoder model to re-rank NRM candidates.
    Much faster than LLM verification and often more accurate
    for short-text classification tasks.
    """
    MODEL_NAME = "cross-encoder/ms-marco-MiniLM-L-6-v2"

    def __init__(self):
        from sentence_transformers import CrossEncoder
        log.info("Loading cross-encoder '%s' ...", self.MODEL_NAME)
        self.model = CrossEncoder(self.MODEL_NAME)
        log.info("Cross-encoder ready")

    def rerank(self, description, candidates):
        """Re-rank candidates using cross-encoder scores."""
        pairs = [(description, f"{code} {desc}") for code, desc, _ in candidates]
        scores = self.model.predict(pairs)
        best_idx = int(np.argmax(scores))
        code, desc, _ = candidates[best_idx]
        confidence = f"{float(scores[best_idx])*100:.1f}%"
        return (code, desc, confidence)


# ══════════════════════════════════════════════════════════════════════
#  LLM VERIFIER (Ollama) - V2: improved prompt with context
# ══════════════════════════════════════════════════════════════════════
class OllamaLLMVerifier:
    DEFAULT_MODEL = "llama3.2:1b"
    DEFAULT_BASE_URL = "http://localhost:11434"

    def __init__(self, model=None, base_url=None):
        self.model = model or self.DEFAULT_MODEL
        self.base_url = (base_url or self.DEFAULT_BASE_URL).rstrip("/")

    def ping(self):
        try:
            resp = requests.get(f"{self.base_url}/api/tags", timeout=5)
            if resp.status_code != 200:
                return False
            models = [m["name"] for m in resp.json().get("models", [])]
            return any(self.model in m for m in models)
        except requests.ConnectionError:
            return False

    def _call_ollama(self, prompt):
        payload = {
            "model": self.model, "prompt": prompt, "stream": False,
            "options": {"temperature": 0.1, "num_predict": 200},
        }
        try:
            resp = requests.post(f"{self.base_url}/api/generate", json=payload, timeout=30)
            resp.raise_for_status()
            return resp.json().get("response", "").strip()
        except Exception as exc:
            log.debug("Ollama call failed: %s", exc)
            return None

    def verify(self, description, candidates, context_hint=None):
        """V2: Enhanced prompt includes context hint about the NRM section."""
        fallback = candidates[0]
        options = "\n".join(
            f"  {i+1}. {code} - {desc} (similarity: {score})"
            for i, (code, desc, score) in enumerate(candidates))

        context_line = ""
        if context_hint:
            context_line = f"\nThis item appears in the '{context_hint}' section of the cost plan.\n"

        prompt = (
            "You are an expert UK construction cost consultant.\n"
            "A cost plan line item is described as:\n"
            f'  "{description}"\n'
            f"{context_line}\n"
            "The following NRM sub-element codes are possible matches:\n"
            f"{options}\n\n"
            "IMPORTANT RULES:\n"
            "- 'External Walls' must map to 2.5.x (External walls), NOT 2.7.x (Internal)\n"
            "- 'Site Works' typically maps to 8.x (External Works)\n"
            "- 'Drainage' without 'internal' or 'disposal' prefix maps to 8.6.x (External drainage)\n"
            "- Do NOT assign NRM codes to fee, contingency, or total rows\n\n"
            "Pick the SINGLE best NRM code. Reply with ONLY a JSON object:\n"
            '{"code": "<nrm_code>", "reason": "<brief reason>"}\n'
        )

        raw = self._call_ollama(prompt)
        if not raw:
            return fallback
        try:
            cleaned = raw.strip().strip("`").strip()
            if cleaned.startswith("json"):
                cleaned = cleaned[4:].strip()
            data = json.loads(cleaned)
            chosen_code = str(data.get("code", "")).strip()
        except (json.JSONDecodeError, AttributeError):
            return fallback

        for code, desc, score in candidates:
            if code == chosen_code:
                return (code, desc, score)
        return fallback


# ══════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════
def write_tables_to_excel(tables_data, output_path):
    LABEL_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")
    GAP_ROWS = 2
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    if not tables_data:
        ws.cell(row=1, column=1, value="No tables were detected in this PDF.")
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=12, italic=True)
        wb.save(output_path)
        return

    current_row = 1
    for t_idx, entry in enumerate(tables_data):
        label = entry["sheet_name"]
        df = entry["dataframe"]
        num_cols = len(df.columns)
        has_nrm = "NRM Code" in df.columns

        label_cell = ws.cell(row=current_row, column=1, value=f"Table {t_idx+1} -- {label}")
        label_cell.font = LABEL_FONT
        current_row += 1

        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        current_row += 1

        for row_tuple in df.itertuples(index=False):
            for c_idx, value in enumerate(row_tuple, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)
                cell.border = THIN_BORDER
                if has_nrm and c_idx > (num_cols - 3):
                    cell.font = NRM_FONT
                    cell.fill = NRM_FILL
                    cell.alignment = CELL_ALIGN
                else:
                    cell.font = CELL_FONT
                    cell.alignment = CELL_ALIGN
            current_row += 1
        current_row += GAP_ROWS

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    note_cell = ws.cell(row=current_row, column=1,
        value="[Extracted via Docling AI (TableFormer ACCURATE) | NRM matching V2 with context-awareness]")
    note_cell.font = Font(name="Calibri", size=9, italic=True, color="888888")
    wb.save(output_path)
    log.info("Saved  %s  (%d table(s))", output_path, len(tables_data))


# ══════════════════════════════════════════════════════════════════════
#  V2: DOCLING CONVERTER FACTORY (tuned pipeline)
# ══════════════════════════════════════════════════════════════════════
def create_converter(use_accurate_mode=True):
    """
    V2: Create a Docling DocumentConverter with tuned pipeline options.
    - TableFormerMode.ACCURATE for better table structure recognition
    - do_cell_matching=True to map structure back to PDF cells
    - OCR enabled for scanned pages
    """
    pipeline_options = PdfPipelineOptions(do_table_structure=True)
    pipeline_options.table_structure_options.mode = (
        TableFormerMode.ACCURATE if use_accurate_mode else TableFormerMode.FAST
    )
    pipeline_options.table_structure_options.do_cell_matching = True
    pipeline_options.do_ocr = True

    converter = DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
        }
    )
    return converter


# ══════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════
def process_pdf(converter, pdf_path, nrm_matcher=None, llm_verifier=None, cross_encoder=None):
    pdf_path = Path(pdf_path)
    output_path = pdf_path.with_name(pdf_path.stem + "_v2_tables.xlsx")
    log.info("Processing  %s", pdf_path.name)

    result = converter.convert(str(pdf_path))
    doc = result.document
    tables = list(doc.tables)
    log.info("  Found %d table(s)", len(tables))

    tables_data = []
    for t_idx, table in enumerate(tables):
        try:
            df = table.export_to_dataframe(doc=doc)
            page_label = ""
            if hasattr(table, "prov") and table.prov:
                try:
                    page_no = table.prov[0].page_no
                    page_label = f"P{page_no}"
                except (IndexError, AttributeError):
                    pass

            sheet_name = f"{page_label}_Table_{t_idx+1}" if page_label else f"Table_{t_idx+1}"

            if nrm_matcher:
                df = nrm_matcher.enrich_dataframe(df, llm_verifier=llm_verifier, cross_encoder=cross_encoder)

            log.info("  Table %d: %d rows x %d cols %s",
                     t_idx+1, len(df), len(df.columns),
                     f"(page {page_label})" if page_label else "")

            tables_data.append({"sheet_name": sheet_name, "dataframe": df})
        except Exception as exc:
            log.warning("  Table %d: export failed - %s", t_idx+1, exc)

    write_tables_to_excel(tables_data, output_path)
    return output_path


def _find_nrm_db(target_path):
    search_dirs = [Path(target_path)]
    if search_dirs[0].is_file():
        search_dirs = [search_dirs[0].parent]
    search_dirs.append(Path.cwd())
    for d in search_dirs:
        candidate = d / "nrm_db.xlsx"
        if candidate.exists():
            return candidate
    return None


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Error: please provide a PDF file or folder path.")
        sys.exit(1)

    target = None
    nrm_db_path = None
    synonyms_path = None
    skip_llm = False
    use_cross_encoder = False
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--nrm-db" and i+1 < len(args):
            nrm_db_path = Path(args[i+1]); i += 2
        elif args[i] == "--synonyms" and i+1 < len(args):
            synonyms_path = Path(args[i+1]); i += 2
        elif args[i] == "--no-llm":
            skip_llm = True; i += 1
        elif args[i] == "--use-cross-encoder":
            use_cross_encoder = True; i += 1
        else:
            target = Path(args[i]); i += 1

    if not target:
        print("Error: please provide a PDF file or folder path.")
        sys.exit(1)

    if target.is_file() and target.suffix.lower() == ".pdf":
        pdf_files = [target]
    elif target.is_dir():
        pdf_files = sorted(target.glob("*.pdf"))
        if not pdf_files:
            log.error("No PDF files found in %s", target); sys.exit(1)
        log.info("Found %d PDF(s) in %s", len(pdf_files), target)
    else:
        log.error("Invalid path: %s", target); sys.exit(1)

    if not nrm_db_path:
        nrm_db_path = _find_nrm_db(target)

    nrm_matcher = None
    if nrm_db_path and nrm_db_path.exists():
        log.info("NRM database found: %s", nrm_db_path)
        nrm_matcher = NRMMatcher(nrm_db_path, synonyms_path=synonyms_path)
    else:
        log.warning("NRM database not found - skipping NRM enrichment.")

    # V2: Cross-encoder re-ranker (optional, faster than LLM)
    cross_enc = None
    if nrm_matcher and use_cross_encoder:
        try:
            cross_enc = CrossEncoderReranker()
        except Exception as exc:
            log.warning("Cross-encoder init failed: %s", exc)

    llm_verifier = None
    if nrm_matcher and not skip_llm and not cross_enc:
        verifier = OllamaLLMVerifier()
        if verifier.ping():
            llm_verifier = verifier
            log.info("Ollama LLM verifier ready (model: %s)", verifier.model)
        else:
            log.warning("Ollama not available - LLM verification disabled.")

    # V2: Use tuned converter
    log.info("Initialising Docling AI models (ACCURATE mode)...")
    converter = create_converter(use_accurate_mode=True)
    log.info("Models loaded.")

    results = []
    for pdf_file in pdf_files:
        try:
            out = process_pdf(converter, pdf_file, nrm_matcher, llm_verifier, cross_enc)
            results.append((pdf_file.name, out, None))
        except Exception as exc:
            log.error("FAILED  %s : %s", pdf_file.name, exc)
            results.append((pdf_file.name, None, str(exc)))

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for name, out, err in results:
        if err:
            print(f"  [FAIL]  {name}  ->  ERROR: {err}")
        else:
            print(f"  [OK]    {name}  ->  {out.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
